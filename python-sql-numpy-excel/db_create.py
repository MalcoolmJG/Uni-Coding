#Script 1 Access the workbooks and create a database
import openpyxl
import sqlite3
#Load Workbooks
wb1 = openpyxl.load_workbook('GlobalLandTemperaturesByCountry.xlsx')
wb2 = openpyxl.load_workbook('GlobalLandTemperaturesByState.xlsx')
wb3 = openpyxl.load_workbook('GlobalLandTemperaturesByMajorCity.xlsx')
#Connect to database
con = sqlite3.connect("database.db")
c = con.cursor()
#Create tables
byCountry_table = """CREATE TABLE temp_by_country(
                     date DATE,
                     aver_temp DECIMAL(5, 3),
                     aver_temp_uncer DECIMAL(5,3),
                     country VARCHAR(50));"""
byState_table = """CREATE TABLE temp_by_state(
                   date DATE,
                   aver_temp DECIMAL(5, 3),
                   aver_temp_uncer DECIMAL(5,3),
                   state VARCHAR(50),
                   country VARCHAR(50));"""
byCity_table = """CREATE TABLE temp_by_city(
                   date DATE,
                   aver_temp DECIMAL(5, 3),
                   aver_temp_uncer DECIMAL(5,3),
                   city VARCHAR(50),
                   country VARCHAR(50),
                   latitude VARCHAR(7),
                   longitude VARCHAR(7));"""
#Dropping table inside the statements doesnt work?
c.execute("DROP TABLE IF EXISTS temp_by_country;")
c.execute(byCountry_table)
c.execute("DROP TABLE IF EXISTS temp_by_state;")
c.execute(byState_table)
c.execute("DROP TABLE IF EXISTS temp_by_city;")
c.execute(byCity_table)
#Insert values into tables
#By Country
insert_into ="""INSERT INTO temp_by_country (date, aver_temp, aver_temp_uncer, country) VALUES(?, ?, ?, ?)"""
sheet = wb1.get_active_sheet()
#loop through excel file rows
for r in range(2, sheet.max_row):
    date = sheet.cell(row = r, column = 1).value
    avTemp = sheet.cell(row = r, column = 2).value
    avTempUncert = sheet.cell(row = r, column = 3).value
    country = sheet.cell(row = r, column = 4).value
    values = (date, avTemp, avTempUncert, country)
    c.execute(insert_into, values)
print('Country done.')
#By State
insert_into="""INSERT INTO temp_by_state (date, aver_temp, aver_temp_uncer, state, country) VALUES(?, ?, ?, ?, ?)"""
sheet = wb2.get_active_sheet()
#loop through excel file rows
for r in range(2, sheet.max_row):
    date = sheet.cell(row = r, column = 1).value
    avTemp = sheet.cell(row = r, column = 2).value
    avTempUncert = sheet.cell(row = r, column = 3).value
    state = sheet.cell(row = r, column = 4).value
    country = sheet.cell(row = r, column = 5).value
    values = (date, avTemp, avTempUncert, state, country)
    c.execute(insert_into, values)
print('State done.')
#By Major City
insert_into="""INSERT INTO temp_by_city (date, aver_temp, aver_temp_uncer, city, country, latitude, longitude) VALUES(?, ?, ?, ?, ?, ?, ?)"""
sheet = wb3.get_active_sheet()
#loop through excel file rows
for r in range(2, sheet.max_row):
    date = sheet.cell(row = r, column = 1).value
    avTemp = sheet.cell(row = r, column = 2).value
    avTempUncert = sheet.cell(row = r, column = 3).value
    city = sheet.cell(row = r, column = 4).value
    country = sheet.cell(row = r, column = 5).value
    lat = sheet.cell(row = r, column = 6).value
    long = sheet.cell(row = r, column = 7).value
    values = (date, avTemp, avTempUncert, city, country, lat, long)
    c.execute(insert_into, values)
print('City done.')
#save
con.commit()
con.close()