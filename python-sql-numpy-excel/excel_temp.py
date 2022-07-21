# -*- coding: utf-8 -*-
#Script 3 Excel via Python
import sqlite3
import openpyxl
from openpyxl.chart import (
    LineChart,
    Reference,)


con = sqlite3.connect("database.db")
c = con.cursor()
sql_command = """SELECT avg(aver_temp), avg(aver_temp_uncer), city
                  FROM temp_by_city
                  WHERE country LIKE 'China'
                  GROUP BY city;"""
wb = openpyxl.Workbook()
sheet = wb.active
#Title sheet and label columns
sheet.title = "Temperature by city"
sheet['A1'] = "Mean Temperature"
sheet['B1'] = "Mean Temperature Uncertainty"
sheet['C1'] = "City"
#find relevant data
c.execute(sql_command)
fetch = c.fetchall()
#write DB into Excel file
r = 2
for row in fetch:
    sheet.cell(row = r, column = 1, value = row[0])
    sheet.cell(row = r, column = 2, value = row[1])
    sheet.cell(row = r, column = 3, value = row[2])    
    r += 1
#Create line graph
chart = LineChart()
chart.title = "Mean temperature in Chinises Cities"
chart.y_axis.title = 'Mean Temperature'
chart.x_axis.title = 'Cities'
datay = Reference(sheet, min_col=1, min_row=1, max_col=sheet.max_column - 1, max_row=sheet.max_row)
chart.add_data(datay, titles_from_data=True)
datax = Reference(sheet, min_col=sheet.max_column, min_row=2, max_col=sheet.max_column, max_row=sheet.max_row)
chart.set_categories(datax)

sheet.add_chart(chart, "E2")
#save
wb.save('World Temperature.xlsx')
con.close()