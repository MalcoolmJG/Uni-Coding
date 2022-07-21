#Script 4 Numpy in Python
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import sqlite3
#Calulate each states Mean temperature by year
def array_loop(input1, input2):
    if input1 == 1:
        sql = """SELECT STRFTIME('%Y', date)
                    FROM temp_by_state
                    WHERE state LIKE ? AND country LIKE "Australia"
                    GROUP BY strftime("%Y", date);"""
        c.execute(sql, (input2,))
        fetch = c.fetchall()
        return fetch
    else:
        sql = """SELECT AVG(aver_temp)
                    FROM temp_by_state
                    WHERE state LIKE ? AND country LIKE "Australia"
                    GROUP BY strftime("%Y", date);"""
        c.execute(sql, (input2,))
        fetch = c.fetchall()
        return fetch
#compare differences 
def diff_loop(time, temp):
    z = []
    x= 0
    for row in time:
        y=0
        for orow in nation_time:
            if orow == row:
                z.append(nation_temp[y] - temp[x])
                y+=1

        x+=1
    return z
#insert into excel file
def xl_insert(title, data, pos):
    sheet.cell(row = 1, column = pos, value = title)
    r = 2
    for roww in data:
        sheet.cell(row = r, column = pos, value = roww[0])
        r+=1    
#Open world temp
wb = openpyxl.load_workbook('World Temperature.xlsx')
wb.create_sheet("Comparison")
sheet = wb.get_sheet_by_name('Comparison')
#conntect to DB
con = sqlite3.connect("database.db")
c = con.cursor()
#Put statements into numpy arrays
act_time = array_loop(1, 'Australian Capital Territory')
act_temp = array_loop(2, 'Australian Capital Territory')
nsw_time = array_loop(1, 'New South Wales')
nsw_temp = array_loop(3, 'New South Wales')
nt_time = array_loop(1, 'Northern Territory')
nt_temp = array_loop(4, 'Northern Territory')
qld_time = array_loop(1, 'Queensland')
qld_temp = array_loop(5, 'Queensland')
sa_time = array_loop(1, 'South Australia')
sa_temp = array_loop(6, 'South Australia')
tas_time = array_loop(1, 'Tasmania')
tas_temp = array_loop(7, 'Tasmania')
vic_time = array_loop(1, 'Victoria')
vic_temp = array_loop(8, 'Victoria')
wa_time = array_loop(1, 'Western Australia')
wa_temp = array_loop(9, 'Western Australia')
#Calulate National Average temperature by year
sql_nation_time = """SELECT STRFTIME('%Y', date)
                FROM temp_by_country
                WHERE country LIKE "Australia"
				GROUP BY strftime("%Y", date);"""
sql_nation_temp = """SELECT aver_temp
                FROM temp_by_country
                WHERE country LIKE "Australia"
				GROUP BY strftime("%Y", date);""" 
c.execute(sql_nation_time)
fetch= c.fetchall()
nation_time = np.array(fetch)
c.execute(sql_nation_temp)
fetch = c.fetchall()
nation_temp = np.array(fetch)

#plot data onto graph
plt.plot(act_time, act_temp, label='ACT')
plt.plot(nsw_time, nsw_temp, label='NSW')
plt.plot(nt_time, nt_temp, label='NT')
plt.plot(qld_time, qld_temp, label='QLD')
plt.plot(sa_time, sa_temp, label='SA')
plt.plot(tas_time, tas_temp, label='TAS') 
plt.plot(vic_time, vic_temp, label='VIC') 
plt.plot(wa_time, wa_temp, label='WA')
plt.plot(nation_time, nation_temp, label='National')
plt.xlabel('Year')
plt.ylabel('Temperature')
plt.title('Australian states temperatures over years')
plt.legend(bbox_to_anchor=(1.05, 1), loc=2, borderaxespad=0.)
plt.figure(figsize=(60,30))

act_diff = diff_loop(act_time, act_temp)
nsw_diff = diff_loop(nsw_time, nsw_temp)
#nt_diff = diff_loop(nt_time, nt_temp)
qld_diff = diff_loop(qld_time, qld_temp)
sa_diff = diff_loop(sa_time, sa_temp)
tas_diff = diff_loop(tas_time, tas_temp)
vic_diff = diff_loop(vic_time, vic_temp)
#wa_diff = diff_loop(wa_time, wa_temp)

xl_insert('ACT Difference', act_diff, 1)
xl_insert('NSW Difference', nsw_diff, 2)
#xl_insert('NT Difference', nt_diff, 3)
xl_insert('QLD Difference', qld_diff, 4)
xl_insert('SA Difference', sa_diff, 5)
xl_insert('TAS Difference', tas_diff, 6)
xl_insert('VIC Difference', vic_diff, 7)
#xl_insert('WA Difference', wa_diff, 8)


#save
wb.save('World Temperature.xlsx')
con.close()